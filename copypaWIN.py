import pyperclip
from openpyxl import Workbook, load_workbook
import pyautogui
import time

# 既存のExcelファイルを読み込むか、新規作成
file_path = 'output.xlsx'
try:
    wb = load_workbook(file_path)
    ws = wb.active
    print(f'Loaded existing workbook {file_path}')
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    print(f'Created new workbook {file_path}')

# 現在のクリップボードの内容を取得
previous_clipboard_content = pyperclip.paste()

def save_to_excel(text):
    row_data = text.split('\t')  # タブ区切りで分割
    ws.append(row_data)
    print(f'Text copied to Excel: {text}')

def paste_to_excel():
    # Excelがアクティブになっていることを確認
    pyautogui.hotkey('alt', 'tab')
    time.sleep(0.5)
    
    # クリップボードの内容をペースト
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.5)
    
    # カーソルを1つ下に移動
    pyautogui.press('down')

# クリップボードの監視ループ
while True:
    try:
        # 現在のクリップボードの内容を取得
        current_clipboard_content = pyperclip.paste()
        
        # クリップボードの内容が変わった場合
        if current_clipboard_content != previous_clipboard_content:
            previous_clipboard_content = current_clipboard_content
            save_to_excel(current_clipboard_content)
            paste_to_excel()
        
        # CPU使用率を抑えるために少し待機
        time.sleep(1)
    
    except KeyboardInterrupt:
        # スクリプト終了時にExcelファイルを保存
        wb.save(file_path)
        print(f'監視を終了しました。Excelファイルを {file_path} に保存しました')
        break
