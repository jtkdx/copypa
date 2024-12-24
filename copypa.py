import pyperclip
from openpyxl import Workbook, load_workbook
import time
from appscript import app

# 既存のExcelファイルを読み込むか、新規作成
file_path = 'output.xlsx'
try:
    wb = load_workbook(file_path)
    ws = wb.active
    print(f'Loaded existing workbook {file_path}')
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    print(f'Created new workbook {file_path}')

# 現在のクリップボードの内容を取得
previous_clipboard_content = pyperclip.paste()

def save_to_excel(text):
    ws.append([text])
    print(f'Text copied to Excel: {text}')

# クリップボードの監視ループ
while True:
    try:
        # 現在のクリップボードの内容を取得
        current_clipboard_content = pyperclip.paste()
        
        # クリップボードの内容が変わった場合
        if current_clipboard_content != previous_clipboard_content:
            previous_clipboard_content = current_clipboard_content
            save_to_excel(current_clipboard_content)
            
            # Excelにテキストをペースト
            excel = app('Microsoft Excel')
            workbook = excel.active_workbook
            sheet = workbook.active_sheet
            
            sheet.range('A1').select()
            excel.activate()
            sheet.cells[excel.selection.row, excel.selection.column].value.set(current_clipboard_content)
            
            # カーソルを1つ下に移動
            excel.selection.offset(1, 0).select()
        
        # CPU使用率を抑えるために少し待機
        time.sleep(1)
    
    except KeyboardInterrupt:
        # スクリプト終了時にExcelファイルを保存
        wb.save(file_path)
        print(f'監視を終了しました。Excelファイルを {file_path} に保存しました')
        break
